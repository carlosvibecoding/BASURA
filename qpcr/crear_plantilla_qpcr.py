#!/usr/bin/env python3
"""
Genera la plantilla Excel para análisis qPCR y opcionalmente valida con un RAW de ejemplo.

Uso:
  python crear_plantilla_qpcr.py
  python crear_plantilla_qpcr.py --demo "PLACA 2 RGS12 060526_data.xls"
"""
from __future__ import annotations

import argparse
import statistics
import sys
from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from processor import (
    SD_THRESHOLD,
    calculate_all,
    default_group_label,
    goi_display_name,
    parse_group_labels,
    parse_raw_rows,
    sample_prefix,
    sample_sort_key,
)

ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "qPCR_plantilla.xlsx"

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL_PPIA = PatternFill("solid", fgColor="D9E1F2")
HEADER_FILL_SYP = PatternFill("solid", fgColor="E2EFDA")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
GREEN_FILL = PatternFill("solid", fgColor="92D050")
RED_FONT = Font(color="FF0000", bold=True)
ORANGE_FONT = Font(color="FF8000", bold=True)
BOLD = Font(bold=True)


def sheet_rows_from_xls(path: Path, sheet_name: str | None = None) -> list:
    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_index(0) if sheet_name is None else wb.sheet_by_name(sheet_name)
    return [sh.row_values(r) for r in range(sh.nrows)]


def write_instructions(ws) -> None:
    ws.title = "Instrucciones"
    ws.column_dimensions["A"].width = 95
    lines = [
        "PLANTILLA qPCR — ΔΔCt (PPIA y SYP)",
        "",
        "1. Pegue el export del termociclador (StepOne) en la hoja RAW, desde la celda A1.",
        "   Puede pegar varias exportaciones una debajo de otra; al pulsar «Procesar» se analiza todo el bloque.",
        "2. Importe la macro (solo la primera vez): Alt+F11 → Archivo → Importar archivo → qpcr/Modulo_qPCR.bas",
        "3. En la hoja RAW, pulse el botón «Procesar placa» (o ejecute la macro ProcesarPlaca).",
        "",
        f"4. Las muestras con Ct SD > {SD_THRESHOLD} en el gen de interés se marcan en rojo (se incluyen en los cálculos).",
        "",
        "Hojas generadas:",
        "  • Datos — tabla leída del RAW (Ct por muestra y gen)",
        "  • Calculos — ΔCt y promedio C con fórmulas Excel (auditable)",
        "  • Resultados — presentación (fórmulas enlazadas a Calculos)",
        "  • GLOBAL — resumen por grupo (controles + cada prefijo de muestra)",
        "",
        "CONFIGURACIÓN DE GRUPOS (editable):",
        "  B21 — Prefijos control para promedio ΔCt (ej: C   o   C,CTRL)",
        "  B22 — Nombres de grupos (ej: C=Controles;S=Suicidas;A=Alcohólicos)",
        "  Si B22 está vacío, se usan nombres automáticos según el prefijo (C, S, A…).",
        "  Las muestras pueden ser cualquier letra(s)+número: C10, S5, A12, ALC3…",
        "  Logo: macro ElegirLogo o figura qPCR_logo en Instrucciones (B24 = ruta).",
        "",
        "Cálculos (por cada gen control PPIA y SYP):",
        "  Paso 1: ΔCt = Ct mean (gen problema) − Ct mean (control)",
        "  Paso 2: Promedio del paso 1 solo en muestras control (prefijo B21)",
        "  Paso 3: ΔΔCt = ΔCt − promedio del paso 2 (fijo para todas las muestras)",
        "  Paso 4: 2^(−ΔΔCt)",
    ]
    for i, line in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)
    ws["A20"] = "CONFIGURACIÓN DE GRUPOS"
    ws["A20"].font = Font(bold=True, size=12)
    ws["A21"] = "Prefijos control (promedio ΔCt):"
    ws["B21"] = "C"
    ws["A22"] = "Nombres grupos (C=Controles;S=Suicidas;A=Alcohólicos):"
    ws["B22"] = ""
    ws["A23"] = "Logo (opcional):"
    ws["A24"] = "Ruta imagen o figura qPCR_logo en esta hoja (macro ElegirLogo):"
    ws["B24"] = ""
    ws.column_dimensions["B"].width = 55


def setup_raw_sheet(ws) -> None:
    ws.title = "RAW"
    ws["A1"] = ""
    ws["A1"].fill = PatternFill("solid", fgColor="EDF2F7")
    for col in range(1, 12):
        ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor="EDF2F7")
    ws["A3"] = "Pegue aqui el export StepOne (Sample Name, Target Name, Ct, Ct Mean, Ct SD...)"
    ws["A3"].font = Font(italic=True, color="666666")
    ws.column_dimensions["A"].width = 14
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 12
    # Nota para botón: se añade con xlsxwriter en build_xlsm_button helper o manualmente
    ws["A2"] = ""


def write_results_headers(ws, goi_label: str) -> None:
    blocks = [
        (1, "PPIA", HEADER_FILL_PPIA),
        (17, "SYP", HEADER_FILL_SYP),
    ]
    headers = [
        "Sample Name",
        "Target Name",
        "Ct",
        "Ct Mean",
        "Ct SD",
        "ΔCt",
        "Prom. ΔCt (C)",
        "ΔΔCt",
        "2^(-ΔΔCt)",
    ]
    for start_col, ref_name, fill in blocks:
        for offset, h in enumerate(headers):
            c = start_col + offset
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = BOLD
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
        title_cell = ws.cell(row=1, column=start_col, value=f"{goi_label} vs {ref_name}")
    ws.freeze_panes = "A2"


def write_sample_block(
    ws,
    row: int,
    start_col: int,
    sample: str,
    target: str,
    ct1,
    ct2,
    ct_mean: float,
    ct_sd: float,
    dct: float,
    dct_mean_c: float,
    ddct: float,
    fc: float,
    flag_red: bool,
    write_calcs: bool,
    flag_indet: bool = False,
    flag_one_rep: bool = False,
    ct1_show: object = None,
    ct2_show: object = None,
    mean_show: object = None,
) -> int:
    """Escribe 2 filas (duplicado). Devuelve la siguiente fila libre."""
    if flag_indet:
        font = RED_FONT
    elif flag_red:
        font = RED_FONT
    elif flag_one_rep:
        font = ORANGE_FONT
    else:
        font = Font()
    r1, r2 = row, row + 1

    def put(r, col, val, calc_col: bool = False):
        if val is None or val == "":
            return
        cell = ws.cell(row=r, column=col, value=val)
        cell.border = BORDER
        if calc_col or col <= start_col + 4:
            cell.font = font

    put(r1, start_col, sample)
    put(r1, start_col + 1, target)
    c1 = ct1_show if ct1_show is not None else ct1
    c2 = ct2_show if ct2_show is not None else ct2
    mshow = mean_show if mean_show is not None else ( "Indeterminado" if flag_indet else ct_mean)
    put(r1, start_col + 2, c1)
    put(r1, start_col + 3, mshow)
    put(r1, start_col + 4, ct_sd)
    if c2 is not None and c2 != "":
        put(r2, start_col + 2, c2)

    if write_calcs:
        if flag_indet:
            put(r1, start_col + 5, "Indeterminado", True)
            put(r1, start_col + 7, "Indeterminado", True)
            put(r1, start_col + 8, "Indeterminado", True)
        else:
            put(r1, start_col + 5, round(dct, 6), True)
            put(r1, start_col + 7, round(ddct, 6), True)
            fc_val = round(fc, 6)
            put(r1, start_col + 8, fc_val, True)
            if fc > 1000 or (fc > 0 and fc < 0.001):
                ws.cell(row=r1, column=start_col + 8).font = RED_FONT

    return row + 2


def write_results_sheet(ws, calcs, goi: str, data) -> None:
    ws.title = "Resultados"
    goi_label = goi_display_name(goi)
    write_results_headers(ws, goi_label)

    # Fila 2: promedio C en columna Prom. dCt (G=7 y W=23)
    avg_ppi = statistics.mean(
        [c.ppi_dct for c in calcs if c.sample.startswith("C") and not c.flag_indeterminate]
    )
    avg_syp = statistics.mean(
        [c.syp_dct for c in calcs if c.sample.startswith("C") and not c.flag_indeterminate]
    )
    ws.cell(row=2, column=1, value="PROMEDIO controles (C)").font = BOLD
    ws.cell(row=2, column=7, value=round(avg_ppi, 6))
    ws.cell(row=2, column=23, value=round(avg_syp, 6))

    row = 3
    for item in calcs:
        r1, r2 = row, row + 1
        fr = item.flag_high_sd or item.flag_indeterminate or item.ppi_fc > 1000
        write_sample_block(
            ws, r1, 1, item.sample, item.goi, item.ct1, item.ct2,
            item.goi_mean, item.goi_sd, item.ppi_dct, item.ppi_dct_mean_c,
            item.ppi_ddct, item.ppi_fc, fr, True,
            item.flag_indeterminate, item.flag_single_rep,
            item.ct1_display, item.ct2_display,
        )
        write_sample_block(
            ws, r1, 17, item.sample, item.goi, item.ct1, item.ct2,
            item.goi_mean, item.goi_sd, item.syp_dct, item.syp_dct_mean_c,
            item.syp_ddct, item.syp_fc, fr, True,
            item.flag_indeterminate, item.flag_single_rep,
            item.ct1_display, item.ct2_display,
        )
        row += 2

    # Genes control: PPIA (izq.) y SYP (der.) en la misma fila por muestra
    for item in calcs:
        ppi = data[item.sample].get("PPIA")
        syp = data[item.sample].get("SYP")
        if ppi:
            cts = ppi.ct_values
            write_sample_block(
                ws, row, 1, item.sample, "PPIA",
                cts[0] if cts else None, cts[1] if len(cts) > 1 else None,
                _mean(ppi) or 0, ppi.ct_sd or 0, 0, 0, 0, 0, False, False,
            )
        if syp:
            cts = syp.ct_values
            write_sample_block(
                ws, row, 17, item.sample, "SYP",
                cts[0] if cts else None, cts[1] if len(cts) > 1 else None,
                _mean(syp) or 0, syp.ct_sd or 0, 0, 0, 0, 0, False, False,
            )
        row += 2


def _mean(reading) -> float | None:
    if reading.ct_mean is not None:
        return reading.ct_mean
    if reading.ct_values:
        return sum(reading.ct_values) / len(reading.ct_values)
    return None


def write_global_sheet(ws, calcs, goi: str, group_labels: dict | None = None) -> None:
    ws.title = "GLOBAL"
    goi_short = goi_display_name(goi)
    labels = group_labels or {}
    control_prefixes = frozenset({"C"})

    def group_title(pref: str) -> str:
        return labels.get(pref, default_group_label(pref))

    controls = sorted(
        [c for c in calcs if sample_prefix(c.sample) in control_prefixes],
        key=lambda x: sample_sort_key(x.sample, control_prefixes),
    )
    by_prefix: dict[str, list] = {}
    for c in calcs:
        pref = sample_prefix(c.sample)
        if pref in control_prefixes:
            continue
        by_prefix.setdefault(pref, []).append(c)

    fills = [GREEN_FILL, PatternFill("solid", fgColor="B4C6E7"), PatternFill("solid", fgColor="FFC000")]

    def write_table(start_col: int, title: str, title_fill, items):
        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=start_col + 3,
        )
        t = ws.cell(row=1, column=start_col, value=title)
        t.font = BOLD
        t.fill = title_fill
        t.alignment = Alignment(horizontal="center")
        for i, h in enumerate(["SUJETO", "PPIA", "SYP", "MEDIA"]):
            c = ws.cell(row=2, column=start_col + i, value=h)
            c.font = BOLD
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")
        for r_idx, item in enumerate(items, 3):
            if item.flag_indeterminate:
                vals = [item.sample, "Indeterminado", "Indeterminado", "Indeterminado"]
            else:
                media = (item.ppi_fc + item.syp_fc) / 2
                vals = [item.sample, item.ppi_fc, item.syp_fc, media]
            for i, val in enumerate(vals):
                cell = ws.cell(row=r_idx, column=start_col + i, value=round(val, 6) if i > 0 and isinstance(val, float) else val)
                cell.border = BORDER
                if item.flag_indeterminate or item.flag_high_sd:
                    cell.font = RED_FONT
                elif item.flag_single_rep:
                    cell.font = ORANGE_FONT

    write_table(1, f"{group_title('C')} {goi_short} PFC", YELLOW_FILL, controls)
    col = 6
    for i, pref in enumerate(sorted(by_prefix.keys())):
        items = sorted(by_prefix[pref], key=lambda x: sample_sort_key(x.sample, control_prefixes))
        fill = fills[i % len(fills)]
        write_table(col, f"{group_title(pref)} {goi_short} PFC", fill, items)
        col += 5

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14


def build_workbook(demo_raw: Path | None = None) -> Workbook:
    wb = Workbook()
    write_instructions(wb.active)
    setup_raw_sheet(wb.create_sheet("RAW"))

    if demo_raw:
        rows = sheet_rows_from_xls(demo_raw)
        goi, data, order = parse_raw_rows(rows)
        calcs = calculate_all(goi, data, order)
        ws_res = wb.create_sheet("Resultados")
        write_results_sheet(ws_res, calcs, goi, data)
        ws_glob = wb.create_sheet("GLOBAL")
        write_global_sheet(ws_glob, calcs, goi)
        # Copiar RAW de demo
        ws_raw = wb["RAW"]
        for r, row in enumerate(rows, 1):
            for c, val in enumerate(row, 1):
                if val != "":
                    ws_raw.cell(row=r, column=c, value=val)
    else:
        wb.create_sheet("Resultados")
        wb.create_sheet("GLOBAL")

    setup_datos_calc_sheets(wb)
    return wb


def setup_datos_calc_sheets(wb) -> None:
    """Cabeceras en hojas intermedias (la macro las rellena al procesar)."""
    if "Datos" not in wb.sheetnames:
        ws_d = wb.create_sheet("Datos")
    else:
        ws_d = wb["Datos"]
    for c, h in enumerate(
        ["Muestra", "Gen", "Ct 1", "Ct 2", "Ct Mean", "Ct SD", "Indet"], 1
    ):
        ws_d.cell(row=1, column=c, value=h).font = BOLD

    if "Calculos" not in wb.sheetnames:
        ws_c = wb.create_sheet("Calculos")
    else:
        ws_c = wb["Calculos"]
    for c, h in enumerate(
        [
            "Muestra",
            "Ct GOI",
            "Ct PPIA",
            "Ct SYP",
            "dCt PPIA",
            "dCt SYP",
            "Gen interes",
            "Prom dCt (C) PPIA",
            "Prom dCt (C) SYP",
        ],
        1,
    ):
        ws_c.cell(row=1, column=c, value=h).font = BOLD


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--demo", type=Path, help="Rellenar con un RAW de ejemplo (.xls)")
    parser.add_argument("-o", "--output", type=Path, default=OUTPUT)
    args = parser.parse_args()

    wb = build_workbook(args.demo)
    wb.save(args.output)
    print(f"Plantilla guardada: {args.output}")
    print("Importe qpcr/Modulo_qPCR.bas en Excel (Alt+F11) y asigne ProcesarPlaca a un botón en RAW.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
