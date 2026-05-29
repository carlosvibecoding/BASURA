#!/usr/bin/env python3
"""
Procesa qPCR SIN macros de Excel.

Uso:
  python procesar_placa.py                           # lee qPCR_plantilla.xlsx hoja RAW
  python procesar_placa.py "PLACA 2 RGS12 060526_data.xls"
  python procesar_placa.py mi_archivo.xlsx -o resultados.xlsx

Requisito: pip install openpyxl xlrd
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    import xlrd
except ImportError:
    xlrd = None

from openpyxl import load_workbook

# Importar lógica del paquete
sys.path.insert(0, str(Path(__file__).resolve().parent))
from crear_plantilla_qpcr import build_workbook, sheet_rows_from_xls, write_global_sheet, write_results_sheet
from processor import calculate_all, parse_raw_rows


def rows_from_xlsx(path: Path, sheet: str = "RAW") -> list:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"No existe la hoja '{sheet}' en {path.name}")
    ws = wb[sheet]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def rows_from_file(path: Path, sheet: str = "RAW") -> list:
    suf = path.suffix.lower()
    if suf == ".xls":
        if xlrd is None:
            raise RuntimeError("Instale xlrd: pip install xlrd")
        return sheet_rows_from_xls(path)
    if suf in (".xlsx", ".xlsm"):
        return rows_from_xlsx(path, sheet)
    raise ValueError(f"Formato no soportado: {suf}")


def process(rows: list, output: Path) -> None:
    goi, data, order = parse_raw_rows(rows)
    calcs = calculate_all(goi, data, order)

    wb = build_workbook()
    # Eliminar hojas vacías y rellenar
    for name in ("Resultados", "GLOBAL"):
        if name in wb.sheetnames:
            del wb[name]
    ws_res = wb.create_sheet("Resultados")
    write_results_sheet(ws_res, calcs, goi, data)
    ws_glob = wb.create_sheet("GLOBAL")
    write_global_sheet(ws_glob, calcs, goi)

    if "RAW" in wb.sheetnames:
        ws_raw = wb["RAW"]
        for r, row in enumerate(rows, 1):
            for c, val in enumerate(row, 1):
                if val is not None and val != "":
                    ws_raw.cell(row=r, column=c, value=val)

    wb.save(output)
    print(f"Listo: {output}")
    print(f"  Gen de interés: {goi}")
    print(f"  Muestras: {len(calcs)}")
    flagged = [c.sample for c in calcs if c.flag_high_sd]
    if flagged:
        print(f"  Ct SD > 0.3 (revisar): {', '.join(flagged)}")


def main() -> int:
    root = Path(__file__).resolve().parent.parent
    parser = argparse.ArgumentParser(description="Procesar qPCR sin macros Excel")
    parser.add_argument(
        "entrada",
        nargs="?",
        default=str(root / "qPCR_plantilla.xlsx"),
        help="Archivo .xls/.xlsx con datos (o plantilla con hoja RAW rellena)",
    )
    parser.add_argument(
        "-o",
        "--salida",
        default=str(root / "qPCR_resultados.xlsx"),
        help="Archivo de salida",
    )
    parser.add_argument(
        "--hoja",
        default="RAW",
        help="Nombre de hoja con el pegado del termociclador (por defecto RAW)",
    )
    args = parser.parse_args()

    entrada = Path(args.entrada)
    if not entrada.exists():
        print(f"No existe: {entrada}", file=sys.stderr)
        return 1

    try:
        if entrada.suffix.lower() == ".xls":
            rows = rows_from_file(entrada)
        else:
            rows = rows_from_file(entrada, args.hoja)
        process(rows, Path(args.salida))
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
